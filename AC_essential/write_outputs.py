#!/usr/bin/env python3
"""Assemble the five data-mining deliverables. Run after build_mining_outputs.py exists."""
import json, os, sys, re, shutil
from collections import defaultdict
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_mining_outputs as B
import openpyxl
from openpyxl.utils import get_column_letter

BASE = B.BASE
S = json.load(open(os.path.join(BASE, 'screened.json')))

def local(u):
    n = u.rstrip('/').split('/')[-1]
    if not n.lower().endswith(('.txt', '.csv', '.tsv')):
        n += '.txt'
    return os.path.join(B.TXT, n)

def safe_tab(name, used):
    t = re.sub(r'[\[\]\*:\?/\\]', '-', str(name))[:31].strip() or 'core'
    base, i = t, 2
    while t.lower() in used:
        suf = f'_{i}'
        t = base[:31 - len(suf)] + suf
        i += 1
    used.add(t.lower())
    return t

# ------------------------------------------------------------------ harvest
flags, rejects, jerry_rows = [], [], []
harvest = defaultdict(list)   # bucket -> list of (meta, series, hdr)

for bucket in ('acc', 'chrono'):
    for r in S[bucket]:
        p = local(r['url'])
        if not os.path.exists(p):
            rejects.append(dict(core=r['site'], url=r['url'], reason='Download missing on disk'))
            continue
        out, err = B.extract(p, r)
        if err:
            tgt = flags if 'HTML' in err or 'Could not locate' in err else rejects
            tgt.append(dict(core=r['site'], url=r['url'], reason=err,
                            study=r.get('study_name', '')))
            continue
        hdr = B.header_text(B.read_lines(p))
        for ser in out:
            harvest[bucket].append((r, ser, hdr))

for r in S['jerry']:
    jerry_rows.append(r)
for r, why in S['flag']:
    flags.append(dict(core=r['site'], url=r['url'], reason=why, study=r.get('study_name', '')))
for r, why in S['rej']:
    rejects.append(dict(core=r['site'], url=r['url'], reason=why, study=r.get('study_name', '')))

# Turney-type interval-average products: reject from sheets, flag for review
INTERVAL = re.compile(r'turney|early lig|maximum annual|interval average|stack', re.I)

PSUF = r'_(UK37|MgCa|TEX86|Foram|Diatom|Radiolaria|Cocc)$'
def norm_id(s):
    return re.sub(r'[^a-z0-9]', '', re.sub(PSUF, '', str(s), flags=re.I).lower())

# --- reference core IDs already held (Hoffman + user's own tabs) -------------
import csv as _csv
REF_NAMES = {}
with open(os.path.join(BASE, 'Hoffman_Metadata.csv')) as fh:
    for row in _csv.DictReader(fh):
        if row['Proxy-Core-ID'] == 'V28-238':
            continue
        REF_NAMES.setdefault(norm_id(row['Proxy-Core-ID']), ('HOFFMAN', row['Proxy-Core-ID']))
_wb0 = openpyxl.load_workbook(os.path.join(BASE, 'SST_Data_Mining.xlsx'), read_only=True)
for _t in _wb0.sheetnames:
    if _t not in ('TEMPLATE', 'Sample'):
        REF_NAMES.setdefault(norm_id(_t), ('MINED', _t))
_wb0.close()

def collapse(items):
    """One series per (site, proxy): keep the richest; report the rest."""
    best, dropped = {}, []
    for meta, ser, hdr in items:
        k = (str(meta['site']).lower(), ser['proxy'])
        if k not in best or len(ser['recs']) > len(best[k][1]['recs']):
            if k in best:
                dropped.append(best[k])
            best[k] = (meta, ser, hdr)
        else:
            dropped.append((meta, ser, hdr))
    return list(best.values()), dropped

# ------------------------------------------------------------------ sheet writer
def write_series(ws, meta, ser, hdr):
    for c, h in enumerate(B.HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    core = re.sub(r'_(UK37|MgCa|TEX86|Foram|Diatom|Radiolaria|Cocc)$', '', str(meta['site']), flags=re.I)
    ws.cell(row=2, column=1, value=core)
    ws.cell(row=2, column=2, value=meta['lon'])
    ws.cell(row=2, column=3, value=meta['lat'])
    ws.cell(row=2, column=4, value=ser['proxy'])
    ws.cell(row=2, column=22, value=(meta.get('cite') or '')[:1500])
    ws.cell(row=2, column=23, value=meta['url'])
    ws.cell(row=2, column=24, value=B.grab(hdr, ['chronolog', 'age model', 'age control', 'stratigraph', 'tie point', 'tuned']))
    ws.cell(row=2, column=25, value=B.grab(hdr, ['method', 'calibration', 'calculated', 'analy', 'proxy value', 'variables']))
    ws.cell(row=2, column=26, value=B.elevation(meta, hdr))
    for k, rec in enumerate(ser['recs']):
        r = 2 + k
        ws.cell(row=r, column=12, value=None)
        ws.cell(row=r, column=13, value=rec['bd18o'])
        ws.cell(row=r, column=14, value=rec['depth'])
        ws.cell(row=r, column=15, value=rec['pval'])
        ws.cell(row=r, column=16, value=rec['sst'])   # P  (== T, no interpolation)
        ws.cell(row=r, column=18, value=rec['age'])   # R
        ws.cell(row=r, column=19, value=rec['lo'])    # S
        ws.cell(row=r, column=20, value=rec['sst'])   # T
        ws.cell(row=r, column=21, value=rec['hi'])    # U

# ------------------------------------------------------------------ 1. main workbook
MAIN = os.path.join(BASE, 'SST_Data_Mining_corrected_2026-07-20.xlsx')
wb = openpyxl.load_workbook(MAIN)
used = {s.lower() for s in wb.sheetnames}
existing = set(wb.sheetnames)

# (a) name-based dedup: a core we already hold is a chronology match, not a new core
_acc = []
for meta, ser, hdr in harvest['acc']:
    n = norm_id(meta['site'])
    if n in REF_NAMES:
        src, oid = REF_NAMES[n]
        meta = dict(meta); meta['dupe'] = [(src, oid)]
        harvest['chrono'].append((meta, ser, hdr))
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason=f'Same core ID as {src} "{oid}" but coordinates fall in a DIFFERENT 1x1 deg cell '
                                 f'({meta["lon"]},{meta["lat"]}) - grid dedup alone missed it; moved to chronology file',
                          study=meta.get('study_name', '')))
    else:
        _acc.append((meta, ser, hdr))
# (b) collapse repeated files for the same site+proxy
harvest['acc'], dropped = collapse(_acc)
for meta, ser, hdr in dropped:
    flags.append(dict(core=meta['site'], url=meta['url'],
                      reason=f'Additional source file for same core+proxy ({ser["proxy"]}, {len(ser["recs"])} pts) '
                             f'not used; richer file kept instead',
                      study=meta.get('study_name', '')))
harvest['chrono'], dropped2 = collapse(harvest['chrono'])
for meta, ser, hdr in dropped2:
    flags.append(dict(core=meta['site'], url=meta['url'],
                      reason=f'Chronology-file duplicate for same core+proxy ({ser["proxy"]}) not written',
                      study=meta.get('study_name', '')))

n_added = 0
for meta, ser, hdr in harvest['acc']:
    if INTERVAL.search(meta.get('study_name', '')):
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason='Interval-average product (not per-sample ages) - NOT added to spreadsheet',
                          study=meta.get('study_name', '')))
        continue
    nm = str(meta['site'])
    if ser['proxy'] and len([1 for m, s2, _ in harvest['acc'] if m['site'] == meta['site']]) > 1:
        nm = f"{meta['site']}_{ser['proxy']}"
    ws = wb.create_sheet(safe_tab(nm, used))
    write_series(ws, meta, ser, hdr)
    n_added += 1
    for note in ser['notes']:
        flags.append(dict(core=meta['site'], url=meta['url'], reason=note,
                          study=meta.get('study_name', ''), tab=ws.title))
    if ser['proxy'] not in ('UK37', 'Foram', 'MgCa', 'Radiolaria', 'Diatom', 'Cocc'):
        flags.append(dict(core=meta['site'], url=meta['url'], tab=ws.title,
                          reason=f"Proxy '{ser['proxy'] or 'UNKNOWN'}' is outside Hoffman's six proxy types",
                          study=meta.get('study_name', '')))
wb.save(MAIN)
print(f'[1] main workbook: +{n_added} tabs  (total {len(wb.sheetnames)}), originals preserved={len(existing & set(wb.sheetnames))}/{len(existing)}')

# ------------------------------------------------------------------ 2. same core, different chronology
wb2 = openpyxl.Workbook(); wb2.remove(wb2.active); used2 = set()
n2 = 0
for meta, ser, hdr in harvest['chrono']:
    nm = f"{meta['site']}_{ser['proxy']}" if ser['proxy'] else str(meta['site'])
    ws = wb2.create_sheet(safe_tab(nm, used2))
    write_series(ws, meta, ser, hdr)
    ws.cell(row=1, column=27, value='Matches existing core (1x1 deg)')
    ws.cell(row=2, column=27, value='; '.join(f'{a}:{b}' for a, b in meta.get('dupe', [])))
    n2 += 1
if not n2:
    wb2.create_sheet('EMPTY')
wb2.save(os.path.join(BASE, 'same_core_different_chronology.xlsx'))
print(f'[2] same_core_different_chronology.xlsx: {n2} tabs')

# ------------------------------------------------------------------ 3. Jerry depth-only
wb3 = openpyxl.Workbook(); wb3.remove(wb3.active); used3 = set()
n3 = 0
for r in jerry_rows:
    p = local(r['url'])
    ws = wb3.create_sheet(safe_tab(str(r['site']), used3))
    for c, h in enumerate(B.HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=r['site'])
    ws.cell(row=2, column=2, value=r['lon'])
    ws.cell(row=2, column=3, value=r['lat'])
    ws.cell(row=2, column=4, value=r.get('proxy'))
    ws.cell(row=2, column=22, value=(r.get('cite') or '')[:1500])
    ws.cell(row=2, column=23, value=r['url'])
    ws.cell(row=2, column=26, value=r.get('elev'))
    if os.path.exists(p):
        L = B.read_lines(p); cols, rows = B.find_table(L)
        if cols:
            role = B.classify(cols)
            di = role.get('depth', [None])[0]
            si = role.get('sst', [None])[0]
            ws.cell(row=2, column=25, value=B.grab(B.header_text(L), ['method', 'calibration']))
            k = 0
            for rr in rows or []:
                dv = B.num(rr[di]) if di is not None and len(rr) > di else None
                sv = B.num(rr[si]) if si is not None and len(rr) > si else None
                if dv is None and sv is None:
                    continue
                ws.cell(row=2 + k, column=14, value=dv)
                ws.cell(row=2 + k, column=16, value=sv)
                ws.cell(row=2 + k, column=20, value=sv)
                k += 1
    n3 += 1
if not n3:
    wb3.create_sheet('EMPTY')
wb3.save(os.path.join(BASE, 'SST_Depth_Only_For_Jerry.xlsx'))
print(f'[3] SST_Depth_Only_For_Jerry.xlsx: {n3} tabs')

json.dump(flags, open(os.path.join(BASE, 'flags.json'), 'w'), indent=1)
json.dump(rejects, open(os.path.join(BASE, 'rejects.json'), 'w'), indent=1)
print(f'[4] flags={len(flags)}  rejects={len(rejects)}')
