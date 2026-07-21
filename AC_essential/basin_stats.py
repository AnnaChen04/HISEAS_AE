#!/usr/bin/env python3
"""Ocean-basin breakdown of every core in the data-mining deliverables.

Basin assignment is geometric, from lon/lat, using conventional limits:
  Southern  lat < -35 in the Indian/Pacific sectors, < -40 in the Atlantic sector
            (approximate position of the Subtropical Front)
  Arctic    lat > 66.5, or the Nordic Seas north of ~66.5
  Atlantic  lon -70..20 (plus the Caribbean/Gulf and the Mediterranean)
  Indian    lon 20..147 with lat < 30
  Pacific   everything else
"""
import csv, os, re, sys
import openpyxl
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))


def basin(lon, lat):
    if lon is None or lat is None:
        return 'Unknown'
    lon = ((float(lon) + 180) % 360) - 180
    lat = float(lat)
    if lat >= 66.5:
        return 'Arctic'
    if lat <= -40:
        return 'Southern'
    if -35 >= lat > -40 and not (-70 <= lon <= 20):
        return 'Southern'
    if -100 <= lon <= 20:
        return 'Atlantic'
    if 20 < lon <= 147 and lat < 30:
        return 'Indian'
    return 'Pacific'


def norm(s):
    return re.sub(r'[^a-z0-9]', '',
                  re.sub(r'_(UK37|MgCa|TEX86|Foram|Diatom|Radiolaria|Cocc)$', '',
                         str(s), flags=re.I).lower())


def scan_xlsx(path, value_col=20, skip=('TEMPLATE', 'Sample')):
    """-> {tab: (core_id, lon, lat, n_points)}"""
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for t in wb.sheetnames:
        if t in skip:
            continue
        rows = list(wb[t].iter_rows(min_row=2, values_only=True))
        if not rows:
            continue
        cid, lon, lat = rows[0][0], rows[0][1], rows[0][2]
        n = sum(1 for r in rows
                if len(r) >= value_col and isinstance(r[value_col - 1], (int, float)))
        out[t] = (cid, lon, lat, n)
    wb.close()
    return out


def summarise(name, data):
    per = defaultdict(lambda: [0, 0])          # basin -> [cores, total points]
    for tab, (cid, lon, lat, n) in data.items():
        b = basin(lon, lat)
        per[b][0] += 1
        per[b][1] += n
    return name, per


if __name__ == '__main__':
    sets = []
    dm = scan_xlsx(os.path.join(BASE, 'SST_Data_Mining_2026-07-21.xlsx'))
    ch = scan_xlsx(os.path.join(BASE, 'same_core_different_chronology.xlsx'))
    je = scan_xlsx(os.path.join(BASE, 'SST_Depth_Only_For_Jerry.xlsx'))
    sets += [summarise('SST_Data_Mining_2026-07-21', dm),
             summarise('same_core_different_chronology', ch),
             summarise('SST_Depth_Only_For_Jerry', je)]

    # Hoffman: one entry per proxy-core; collapse to unique physical cores too
    hof = {}
    hof_cores = {}
    with open(os.path.join(BASE, 'Hoffman_Metadata.csv')) as fh:
        for row in csv.DictReader(fh):
            k = row['Proxy-Core-ID']
            hof[k] = (k, float(row['Longitude']), float(row['Latitude']), 0)
            hof_cores.setdefault(norm(k), (k, float(row['Longitude']), float(row['Latitude']), 0))
    # Point counts come from the harmonized workbook, which carries the actual
    # time series; the metadata CSV holds coordinates only. Coordinates agree
    # between the two, so either can be used for the basin assignment.
    NON_CORE = ('Fig1 Global & Regional Stacks', 'Reference Cores Tie Points',
                'HadSST & SynTrACE Bias Estimate', 'TEMPLATE', 'NATL Data Read Me',
                'PAC Data Read Me', 'IND Data Read Me', 'SATL Data Read Me')
    HH = os.path.join(BASE, 'SST_Hoffman_Harmonized_AC_ES_no_V28_238.xlsx')
    # col T (20) is the harmonised 0.1 ka series - a constant 151 steps per core by
    # construction. Col P (16) is the RAW proxy SST, i.e. one value per measured
    # sample, which is the like-for-like comparison with the newly mined records.
    sets.append(summarise('SST_Hoffman_Harmonized - interpolated 0.1 ka series (col T)',
                          scan_xlsx(HH, value_col=20, skip=NON_CORE)))
    sets.append(summarise('SST_Hoffman_Harmonized - RAW proxy samples (col P)',
                          scan_xlsx(HH, value_col=16, skip=NON_CORE)))

    BAS = ['Arctic', 'Atlantic', 'Pacific', 'Indian', 'Southern', 'Unknown']
    for name, per in sets:
        tot_c = sum(v[0] for v in per.values())
        tot_p = sum(v[1] for v in per.values())
        print(f'\n### {name}')
        print(f'{"basin":10s} {"cores":>6s} {"points":>8s} {"mean pts/core":>14s}')
        for b in BAS:
            if b not in per:
                continue
            c, p = per[b]
            print(f'{b:10s} {c:6d} {p:8d} {p/c:14.1f}')
        print(f'{"TOTAL":10s} {tot_c:6d} {tot_p:8d} ' +
              (f'{tot_p/tot_c:14.1f}' if tot_c else ''))
    print(f'\nHoffman unique physical cores: {len(hof_cores)} (from {len(hof)} proxy-core rows)')
    ph = defaultdict(int)
    for k, (cid, lon, lat, n) in hof_cores.items():
        ph[basin(lon, lat)] += 1
    print('  by basin:', dict(ph))
