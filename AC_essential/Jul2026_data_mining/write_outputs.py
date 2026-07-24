#!/usr/bin/env python3
"""Assemble the five data-mining deliverables.

Extraction goes through extract2.route (special_formats -> manual_specs -> parser2).
build_mining_outputs is retained ONLY for its static helpers (HEADERS, grab,
elevation); its v1 extract() is no longer called anywhere.
"""
import json, os, sys, re, shutil
from collections import defaultdict
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_mining_outputs as B
import extract2 as X
import sheet_lib as SL
import openpyxl
from openpyxl.utils import get_column_letter

BASE = B.BASE
S = json.load(open(os.path.join(BASE, 'screened.json')))

# Layout + per-core amendment helpers live in sheet_lib so patch_tabs.py can
# regenerate a single tab identically without re-running the whole pipeline.
HEADERS_EXT, COL_SEASON, COL_DUPE = SL.HEADERS_EXT, SL.COL_SEASON, SL.COL_DUPE
apply_fixes, write_series, season_label = SL.apply_fixes, SL.write_series, SL.season_label

# Col AA (27), added by AC 2026-07-21. Row 2 only: 'aug-feb averaged' for cores
# whose annual SST is the mean of paired August/February columns, 'summer signal'
# for cores where only a warm-season column exists. Per-timestep seasonal values
# are deliberately NOT carried into the sheet.

def local(u):
    n = u.rstrip('/').split('/')[-1]
    if not n.lower().endswith(('.txt', '.csv', '.tsv')):
        n += '.txt'
    return os.path.join(B.TXT, n)

def safe_tab(name, used):
    t = re.sub(r'[\[\]\*:\?/\\]', '-', str(name))[:31].strip() or 'core'
    base, i = t, 2
    while t.lower() in used:
        suf = f'_{i}'
        t = base[:31 - len(suf)] + suf
        i += 1
    used.add(t.lower())
    return t

# ------------------------------------------------------------------ harvest
flags, rejects, jerry_rows = [], [], []
harvest = defaultdict(list)   # bucket -> list of (meta, series, hdr)

import special_formats as _SPmod
import decisions as D
_SPECIAL_FILES = set(_SPmod.SPECIAL) | set(_SPmod.SPECIAL_REJECT)


for bucket in ('acc', 'chrono'):
    for r in S[bucket]:
        p = local(r['url'])
        # section 6.3 files are routed once, explicitly, further down - skip here
        # so they are not harvested twice under two different buckets
        if os.path.basename(p) in _SPECIAL_FILES:
            continue
        why = D.excluded(r['site'])
        if why:
            rejects.append(dict(core=r['site'], url=r['url'], reason=why,
                                study=r.get('study_name', ''), parser='AC-decision'))
            continue
        if not os.path.exists(p):
            rejects.append(dict(core=r['site'], url=r['url'], reason='Download missing on disk'))
            continue
        _fx = D.fix_for(r['site'])
        if _fx.get('sst_col'):
            r = dict(r, _force_sst_col=_fx['sst_col'])
        out, err = X.route(p, r)
        if err:
            tgt = flags if 'HTML' in err or 'Could not locate' in err else rejects
            tgt.append(dict(core=r['site'], url=r['url'], reason=err,
                            study=r.get('study_name', ''), parser=X.parser_used(p)))
            continue
        hdr = X.header_text(X.read_lines(p))
        for ser in out:
            ser.setdefault('parser', X.parser_used(p))
            apply_fixes(r, ser)
            # a stack file (herbert2016) yields one series per constituent core
            if ser.get('core_override'):
                r2 = dict(r)
                r2['site'] = ser['core_override']
                if ser.get('lon') is not None:
                    r2['lon'], r2['lat'] = ser['lon'], ser['lat']
                harvest[bucket].append((r2, ser, hdr))
            else:
                harvest[bucket].append((r, ser, hdr))

# ---- section 6.3 special-format files -------------------------------------
# These are NOT all in the acc/chrono buckets - herbert2016-med.txt and
# herbert2016-odp883_884.txt were screened into 'flag', and the hand-edited
# tripati .xlsx has no NCEI url at all. Route them explicitly so a screening
# bucket cannot silently drop a file section 6.3 names.
import special_formats as SP

_meta_by_file = {}
_meta_by_stem = {}
for _b in ('acc', 'chrono', 'jerry', 'flag', 'rej'):
    for _it in S.get(_b, []):
        _r = _it[0] if isinstance(_it, list) else _it
        _meta_by_file.setdefault(os.path.basename(local(_r['url'])), _r)
        # Some screened rows point at a DIRECTORY url (e.g. .../ikehara2000/),
        # so their basename is the study stem, not the data filename. Keep a
        # stem index so 'ikehara2000-tsp-2mc.txt' can still find its metadata -
        # without it the core is written with no coordinates.
        _meta_by_stem.setdefault(os.path.splitext(_r['url'].rstrip('/').split('/')[-1])[0], _r)


def meta_for(fn):
    if fn in _meta_by_file:
        return _meta_by_file[fn]
    for _stem, _r in _meta_by_stem.items():
        if _stem and fn.lower().startswith(_stem.lower()):
            return _r
    return None

_special_handled = set()
for _fn in list(SP.SPECIAL) + list(SP.SPECIAL_REJECT):
    if _fn in _special_handled:
        continue
    _special_handled.add(_fn)
    _p = os.path.join(B.TXT, _fn)
    _m = dict(meta_for(_fn) or
              dict(site=os.path.splitext(_fn)[0], lon=None, lat=None, elev=None,
                   url=_fn, cite='', study_name='', proxy=''))
    if not os.path.exists(_p):
        rejects.append(dict(core=_m['site'], url=_m['url'],
                            reason=f'Section 6.3 file {_fn} not found on disk'))
        continue
    _out, _err = SP.extract_special(_p, _m)
    if _err:
        rejects.append(dict(core=_m['site'], url=_m['url'], reason=_err,
                            study=_m.get('study_name', ''), parser='special'))
        continue
    _hdr = X.header_text(X.read_lines(_p)) if _fn.lower().endswith(('.txt', '.csv', '.tsv')) else ''
    for _ser in _out:
        _ser.setdefault('parser', 'special')
        _r2 = dict(_m)
        if _ser.get('core_override'):
            _r2['site'] = _ser['core_override']
        for _k in ('lon', 'lat', 'elev'):
            if _ser.get(_k) is not None:
                _r2[_k] = _ser[_k]
        harvest['acc'].append((_r2, _ser, _hdr))

for r in S['jerry']:
    jerry_rows.append(r)
for r, why in S['flag']:
    flags.append(dict(core=r['site'], url=r['url'], reason=why, study=r.get('study_name', '')))
for r, why in S['rej']:
    rejects.append(dict(core=r['site'], url=r['url'], reason=why, study=r.get('study_name', '')))

# Turney-type interval-average products: reject from sheets, flag for review
INTERVAL = re.compile(r'turney|early lig|maximum annual|interval average|stack', re.I)

PSUF = r'_(UK37|MgCa|TEX86|Foram|Diatom|Radiolaria|Cocc)$'
def norm_id(s):
    return re.sub(r'[^a-z0-9]', '', re.sub(PSUF, '', str(s), flags=re.I).lower())

# --- reference core IDs already held (Hoffman + user's own tabs) -------------
import csv as _csv
REF_NAMES = {}
with open(os.path.join(BASE, 'Hoffman_Metadata.csv')) as fh:
    for row in _csv.DictReader(fh):
        if row['Proxy-Core-ID'] == 'V28-238':
            continue
        REF_NAMES.setdefault(norm_id(row['Proxy-Core-ID']), ('HOFFMAN', row['Proxy-Core-ID']))
_wb0 = openpyxl.load_workbook(os.path.join(BASE, 'SST_Data_Mining.xlsx'), read_only=True)
for _t in _wb0.sheetnames:
    if _t not in ('TEMPLATE', 'Sample'):
        REF_NAMES.setdefault(norm_id(_t), ('MINED', _t))
_wb0.close()

def collapse(items):
    """One series per (site, proxy): keep the richest; report the rest."""
    best, dropped = {}, []
    for meta, ser, hdr in items:
        k = (str(meta['site']).lower(), ser['proxy'])
        if k not in best or len(ser['recs']) > len(best[k][1]['recs']):
            if k in best:
                dropped.append(best[k])
            best[k] = (meta, ser, hdr)
        else:
            dropped.append((meta, ser, hdr))
    return list(best.values()), dropped

# ------------------------------------------------------------------ sheet writer

# ------------------------------------------------------------------ 1. main workbook
MAIN = os.path.join(BASE, 'SST_Data_Mining_corrected_2026-07-20.xlsx')
# IDEMPOTENCE. .bak_corrected.xlsx is Anna's 24 tabs (+ TEMPLATE, Sample) with the
# 20 corrections of corrections_log.json applied and NO mined tabs. Rebuilding from
# it every run means re-running this script replaces the mined tabs rather than
# appending a second copy of them. SST_Data_Mining.xlsx is never opened for writing.
BASEBOOK = os.path.join(BASE, '.bak_corrected.xlsx')
if not os.path.exists(BASEBOOK):
    raise SystemExit('.bak_corrected.xlsx missing - cannot rebuild the main workbook safely')

# PRESERVE AC'S HAND EDITS. Rebuilding from .bak_corrected.xlsx would otherwise
# silently discard any tab Anna has fixed by hand since the last run - it did
# exactly that to her IODP U1482 and IODP U1485 corrections on 2026-07-21. Any
# tab whose Method column (Y) says "manually edited" is copied across verbatim
# and is NOT regenerated. Mark a tab that way and the pipeline will leave it alone.
MANUAL_MARK = re.compile(r'manual(ly)?\s*edit', re.I)
_preserved = {}
if os.path.exists(MAIN):
    _prev = openpyxl.load_workbook(MAIN)
    for _t in _prev.sheetnames:
        _y = _prev[_t].cell(row=2, column=25).value
        if _y and MANUAL_MARK.search(str(_y)):
            _preserved[_t] = [[c.value for c in row] for row in _prev[_t].iter_rows()]
    _prev.close()
    if _preserved:
        shutil.copyfile(MAIN, os.path.join(BASE, '.bak_before_rebuild.xlsx'))

shutil.copyfile(BASEBOOK, MAIN)
wb = openpyxl.load_workbook(MAIN)

# core IDs held by a preserved tab, so the harvest does not write a rival copy
_preserved_ids = set()
for _t, _grid in _preserved.items():
    if len(_grid) > 1 and _grid[1]:
        _preserved_ids.add(re.sub(r'[^a-z0-9]', '', str(_grid[1][0]).lower()))
used = {s.lower() for s in wb.sheetnames}
existing = set(wb.sheetnames)

# (a) name-based dedup: a core we already hold is a chronology match, not a new core
_acc = []
for meta, ser, hdr in harvest['acc']:
    n = norm_id(meta['site'])
    if n in REF_NAMES:
        src, oid = REF_NAMES[n]
        meta = dict(meta); meta['dupe'] = [(src, oid)]
        harvest['chrono'].append((meta, ser, hdr))
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason=f'Same core ID as {src} "{oid}" but coordinates fall in a DIFFERENT 1x1 deg cell '
                                 f'({meta["lon"]},{meta["lat"]}) - grid dedup alone missed it; moved to chronology file',
                          study=meta.get('study_name', '')))
    else:
        _acc.append((meta, ser, hdr))
# (b) collapse repeated files for the same site+proxy
harvest['acc'], dropped = collapse(_acc)
for meta, ser, hdr in dropped:
    flags.append(dict(core=meta['site'], url=meta['url'],
                      reason=f'Additional source file for same core+proxy ({ser["proxy"]}, {len(ser["recs"])} pts) '
                             f'not used; richer file kept instead',
                      study=meta.get('study_name', '')))
harvest['chrono'], dropped2 = collapse(harvest['chrono'])
for meta, ser, hdr in dropped2:
    flags.append(dict(core=meta['site'], url=meta['url'],
                      reason=f'Chronology-file duplicate for same core+proxy ({ser["proxy"]}) not written',
                      study=meta.get('study_name', '')))

n_added = 0
for meta, ser, hdr in harvest['acc']:
    # The INTERVAL screen matches on study TITLE, which over-fires: IODP U1485's
    # study is titled "...Regional SST Stacks and Western Pacific Mg/Ca SST", but
    # the file we use is the per-sample Mg/Ca record, not the stack. AC confirmed
    # U1485 is not an interval-average product, so an explicit decision entry
    # overrides the title screen.
    if INTERVAL.search(meta.get('study_name', '')) and not D.fix_for(meta['site']):
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason='Interval-average product (not per-sample ages) - NOT added to spreadsheet',
                          study=meta.get('study_name', '')))
        continue
    if re.sub(r'[^a-z0-9]', '', str(meta['site']).lower()) in _preserved_ids:
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason='Tab is marked "manually edited" in column Y and was preserved '
                                 'verbatim; the pipeline did NOT regenerate it',
                          study=meta.get('study_name', '')))
        continue
    nm = str(meta['site'])
    if ser['proxy'] and len([1 for m, s2, _ in harvest['acc'] if m['site'] == meta['site']]) > 1:
        nm = f"{meta['site']}_{ser['proxy']}"
    ws = wb.create_sheet(safe_tab(nm, used))
    write_series(ws, meta, ser, hdr)
    n_added += 1
    for note in ser['notes']:
        flags.append(dict(core=meta['site'], url=meta['url'], reason=note,
                          study=meta.get('study_name', ''), tab=ws.title))
    if ser['proxy'] not in ('UK37', 'Foram', 'MgCa', 'Radiolaria', 'Diatom', 'Cocc'):
        flags.append(dict(core=meta['site'], url=meta['url'], tab=ws.title,
                          reason=f"Proxy '{ser['proxy'] or 'UNKNOWN'}' is outside Hoffman's six proxy types",
                          study=meta.get('study_name', '')))
for _t, _grid in _preserved.items():
    _ws = wb[_t] if _t in wb.sheetnames else wb.create_sheet(safe_tab(_t, used))
    for _ri, _row in enumerate(_grid, 1):
        for _ci, _v in enumerate(_row, 1):
            _ws.cell(row=_ri, column=_ci, value=_v)

wb.save(MAIN)
print(f'[1] main workbook: +{n_added} tabs  (total {len(wb.sheetnames)}), '
      f'originals preserved={len(existing & set(wb.sheetnames))}/{len(existing)}, '
      f'hand-edited tabs preserved={len(_preserved)} {sorted(_preserved) if _preserved else ""}')

# ------------------------------------------------------------------ 2. same core, different chronology
wb2 = openpyxl.Workbook(); wb2.remove(wb2.active); used2 = set()
n2 = 0
for meta, ser, hdr in harvest['chrono']:
    # AC confirmed the versions of these already in her own spreadsheet are correct,
    # so the chronology duplicates are not written out (handoff section 6).
    if D.norm(meta['site']) in D.DROP_FROM_CHRONO:
        flags.append(dict(core=meta['site'], url=meta['url'],
                          reason='Chronology-file tab dropped: AC confirmed the version already '
                                 'in SST_Data_Mining.xlsx is the correct one',
                          study=meta.get('study_name', '')))
        continue
    nm = f"{meta['site']}_{ser['proxy']}" if ser['proxy'] else str(meta['site'])
    ws = wb2.create_sheet(safe_tab(nm, used2))
    write_series(ws, meta, ser, hdr)
    ws.cell(row=1, column=COL_DUPE, value='Matches existing core (1x1 deg)')
    ws.cell(row=2, column=COL_DUPE, value='; '.join(f'{a}:{b}' for a, b in meta.get('dupe', [])))
    n2 += 1
if not n2:
    wb2.create_sheet('EMPTY')
wb2.save(os.path.join(BASE, 'same_core_different_chronology.xlsx'))
print(f'[2] same_core_different_chronology.xlsx: {n2} tabs')

# ------------------------------------------------------------------ 3. Jerry depth-only
wb3 = openpyxl.Workbook(); wb3.remove(wb3.active); used3 = set()
n3 = 0
for r in jerry_rows:
    p = local(r['url'])
    ws = wb3.create_sheet(safe_tab(str(r['site']), used3))
    for c, h in enumerate(HEADERS_EXT, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=r['site'])
    ws.cell(row=2, column=2, value=r['lon'])
    ws.cell(row=2, column=3, value=r['lat'])
    ws.cell(row=2, column=4, value=r.get('proxy'))
    ws.cell(row=2, column=22, value=(r.get('cite') or '')[:1500])
    ws.cell(row=2, column=23, value=r['url'])
    ws.cell(row=2, column=26, value=r.get('elev'))
    if os.path.exists(p):
        L = B.read_lines(p); cols, rows = B.find_table(L)
        if cols:
            role = B.classify(cols)
            di = role.get('depth', [None])[0]
            si = role.get('sst', [None])[0]
            ws.cell(row=2, column=25, value=B.grab(B.header_text(L), ['method', 'calibration']))
            k = 0
            for rr in rows or []:
                dv = B.num(rr[di]) if di is not None and len(rr) > di else None
                sv = B.num(rr[si]) if si is not None and len(rr) > si else None
                if dv is None and sv is None:
                    continue
                ws.cell(row=2 + k, column=14, value=dv)
                ws.cell(row=2 + k, column=16, value=sv)
                ws.cell(row=2 + k, column=20, value=sv)
                k += 1
    n3 += 1

# --- section 6.3 Ruddiman fossil-plankton set + section 6.4 dense-depth cores -
import jerry_extra as JE

HDR_WARM = COL_SEASON + 1          # AB
HDR_COLD = COL_SEASON + 2          # AC

def write_depth_core(c, src_label):
    ws = wb3.create_sheet(safe_tab(c['core'], used3))
    for col, h in enumerate(HEADERS_EXT, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=1, column=HDR_WARM, value='SSTwarm')
    ws.cell(row=1, column=HDR_COLD, value='SSTcold')
    ws.cell(row=2, column=1, value=c['core'])
    ws.cell(row=2, column=2, value=c['lon'])
    ws.cell(row=2, column=3, value=c['lat'])
    ws.cell(row=2, column=4, value=c['proxy'])
    ws.cell(row=2, column=23, value=src_label)
    ws.cell(row=2, column=25, value=(c['method'] or '')[:2000])
    ws.cell(row=2, column=26, value=c['elev'])
    ws.cell(row=2, column=COL_SEASON, value=c.get('season'))
    for k, rr in enumerate(c['rows']):
        r = 2 + k
        ws.cell(row=r, column=14, value=rr['depth'])
        ws.cell(row=r, column=16, value=rr['sst'])
        ws.cell(row=r, column=20, value=rr['sst'])
        ws.cell(row=r, column=HDR_WARM, value=rr.get('warm'))
        ws.cell(row=r, column=HDR_COLD, value=rr.get('cold'))
    return ws

for c in JE.ruddiman_cores():
    write_depth_core(c, c['src'])
    n3 += 1
for c in JE.dense_depth_cores(_meta_by_file):
    write_depth_core(c, c['src'])
    n3 += 1

if not n3:
    wb3.create_sheet('EMPTY')
wb3.save(os.path.join(BASE, 'SST_Depth_Only_For_Jerry.xlsx'))
print(f'[3] SST_Depth_Only_For_Jerry.xlsx: {n3} tabs')

json.dump(flags, open(os.path.join(BASE, 'flags.json'), 'w'), indent=1)
json.dump(rejects, open(os.path.join(BASE, 'rejects.json'), 'w'), indent=1)
print(f'[4] flags={len(flags)}  rejects={len(rejects)}')
