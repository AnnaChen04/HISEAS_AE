#!/usr/bin/env python3
"""
Surgically regenerate named tabs in the main workbook, leaving every other tab -
including Anna's hand edits - byte-for-byte alone.

Use this instead of a full write_outputs.py rebuild when only a few cores need
correcting and the workbook contains manual edits that must not be disturbed.

    python3 patch_tabs.py "MD98-2152" "VM19-193"
"""
import json, os, re, sys, shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import build_mining_outputs as B
import extract2 as X
import decisions as D
import sheet_lib as W

BASE = B.BASE
MAIN = os.path.join(BASE, 'SST_Data_Mining_corrected_2026-07-20.xlsx')
S = json.load(open(os.path.join(BASE, 'screened.json')))

targets = sys.argv[1:]
if not targets:
    raise SystemExit(__doc__)

def local(u):
    n = u.rstrip('/').split('/')[-1]
    if not n.lower().endswith(('.txt', '.csv', '.tsv')):
        n += '.txt'
    return os.path.join(B.TXT, n)

shutil.copyfile(MAIN, MAIN + '.bak')
wb = openpyxl.load_workbook(MAIN)

for want in targets:
    hit = None
    for bucket in ('acc', 'chrono'):
        for r in S[bucket]:
            if str(r['site']) != want:
                continue
            p = local(r['url'])
            if not os.path.exists(p):
                continue
            fx = D.fix_for(r['site'])
            r2 = dict(r, **({'_force_sst_col': fx['sst_col']} if fx.get('sst_col') else {}))
            out, err = X.route(p, r2)
            if err or not out:
                continue
            hdr = X.header_text(X.read_lines(p))
            for ser in out:
                W.apply_fixes(r, ser)
                hit = (r, ser, hdr)
                tab = want if len(out) == 1 else f"{want}_{ser['proxy']}"
                if tab not in wb.sheetnames:
                    print(f'  ! {tab} not present in workbook - creating')
                    wb.create_sheet(tab)
                ws = wb[tab]
                for row in ws.iter_rows():
                    for c in row:
                        c.value = None
                W.write_series(ws, r, ser, hdr)
                n = len(ser['recs'])
                mean = sum(x['sst'] for x in ser['recs']) / n
                print(f'  patched {tab:22s} col={ser["col"]:14s} n={n:3d} mean={mean:.2f}')
    if hit is None:
        print(f'  ! {want}: no source series found')

wb.save(MAIN)
print(f'saved {os.path.basename(MAIN)} (previous copy at {os.path.basename(MAIN)}.bak)')
